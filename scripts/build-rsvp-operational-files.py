#!/usr/bin/env python
"""Build final RSVP seed and first-lot dispatch CSVs from the reviewed XLSX."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SOURCE_XLSX = Path("output/convites/convidados-triagem.xlsx")
OUT_DIR = Path("output/convites")
HOST = "https://site-convidados-main.vercel.app"

SEED_COLUMNS = [
    "id_familia",
    "nome_familia",
    "telefone_whatsapp",
    "nome_convidado",
    "e_crianca",
    "e_responsavel_familia",
    "email",
    "observacoes",
]

DISPATCH_COLUMNS = [
    "id_familia",
    "nome_familia",
    "telefone_whatsapp",
    "nomes_grupo",
    "responsavel",
    "codigo_rsvp",
    "link_rsvp",
    "status_envio",
    "observacoes",
]

PENDING_COLUMNS = [
    "id_familia",
    "nome_familia",
    "origem",
    "nomes_grupo",
    "responsavel",
    "telefones",
    "status_pendencia",
    "observacoes",
]


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_yes_no(value: Any, default: str = "nao") -> str:
    text = clean(value).lower()
    return "sim" if text == "sim" else default


def normalize_phone(value: Any) -> str:
    text = clean(value)
    if not text:
        return ""
    prefix = "+" if text.lstrip().startswith("+") else ""
    digits = re.sub(r"\D+", "", text)
    return f"{prefix}{digits}" if prefix else digits


def is_valid_phone(phone: str) -> bool:
    return bool(re.fullmatch(r"\+\d{8,15}", phone))


def read_xlsx(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    raw_rows = list(sheet.iter_rows(values_only=True))
    if not raw_rows:
        raise ValueError(f"Empty workbook: {path}")

    headers = [clean(cell) for cell in raw_rows[0]]
    required = {
        "origem",
        "nome_convidado",
        "id_familia_sugerido",
        "nome_familia_sugerido",
        "telefone_whatsapp",
        "e_crianca",
        "e_responsavel_familia",
        "status_revisao",
        "observacoes",
    }
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(f"Missing XLSX column(s): {', '.join(missing)}")

    rows: list[dict[str, str]] = []
    for raw in raw_rows[1:]:
        if not any(clean(cell) for cell in raw):
            continue
        row = {header: clean(raw[index]) if index < len(raw) else "" for index, header in enumerate(headers)}
        row["telefone_whatsapp"] = normalize_phone(row.get("telefone_whatsapp"))
        row["e_crianca"] = normalize_yes_no(row.get("e_crianca"))
        row["e_responsavel_familia"] = normalize_yes_no(row.get("e_responsavel_familia"))
        rows.append(row)
    return rows


def group_rows(rows: list[dict[str, str]]) -> dict[str, list[dict[str, str]]]:
    groups: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        group_id = row["id_familia_sugerido"]
        if not group_id:
            raise ValueError(f"Guest without id_familia_sugerido: {row.get('nome_convidado')}")
        groups[group_id].append(row)
    return dict(groups)


def is_extra(row: dict[str, str]) -> bool:
    return "LISTA EXTRA" in row.get("origem", "").upper()


def family_name(group: list[dict[str, str]]) -> str:
    responsible = next((row for row in group if row["e_responsavel_familia"] == "sim"), group[0])
    return responsible.get("nome_familia_sugerido") or group[0].get("nome_familia_sugerido") or group[0]["id_familia_sugerido"]


def responsible_name(group: list[dict[str, str]]) -> str:
    responsible = next((row for row in group if row["e_responsavel_familia"] == "sim"), None)
    return responsible.get("nome_convidado", "") if responsible else ""


def group_phones(group: list[dict[str, str]]) -> list[str]:
    seen: set[str] = set()
    phones: list[str] = []
    for row in group:
        phone = row.get("telefone_whatsapp", "")
        if phone and phone not in seen:
            seen.add(phone)
            phones.append(phone)
    return phones


def load_codes(path: Path | None) -> dict[str, str]:
    if not path:
        return {}
    if not path.exists():
        raise FileNotFoundError(path)

    if path.suffix.lower() == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            return {str(key): str(value) for key, value in data.items() if value}
        if isinstance(data, list):
            out: dict[str, str] = {}
            for item in data:
                if isinstance(item, dict):
                    group_id = clean(item.get("id_familia") or item.get("familyId"))
                    code = clean(item.get("codigo_rsvp") or item.get("code"))
                    if group_id and code:
                        out[group_id] = code
            return out
        raise ValueError("Unsupported JSON code map format")

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        out = {}
        for row in reader:
            group_id = clean(row.get("id_familia") or row.get("familyId"))
            code = clean(row.get("codigo_rsvp") or row.get("code"))
            if group_id and code:
                out[group_id] = code
        return out


def write_csv(path: Path, rows: list[dict[str, str]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def build_outputs(rows: list[dict[str, str]], codes: dict[str, str], host: str) -> tuple[list[dict[str, str]], list[dict[str, str]], list[dict[str, str]], list[str]]:
    first_lot_groups = group_rows([row for row in rows if not is_extra(row)])
    extra_groups = group_rows([row for row in rows if is_extra(row)])
    errors: list[str] = []
    seed_rows: list[dict[str, str]] = []
    dispatch_rows: list[dict[str, str]] = []
    pending_rows: list[dict[str, str]] = []

    for group_id, group in first_lot_groups.items():
        phones = group_phones(group)
        invalid = [phone for phone in phones if not is_valid_phone(phone)]
        responsible = [row for row in group if row["e_responsavel_familia"] == "sim"]
        if invalid:
            errors.append(f"{group_id}: invalid phone(s): {', '.join(invalid)}")
        if not phones:
            pending_rows.append({
                "id_familia": group_id,
                "nome_familia": family_name(group),
                "origem": ";".join(sorted({row.get("origem", "") for row in group if row.get("origem")})),
                "nomes_grupo": " | ".join(row["nome_convidado"] for row in group),
                "responsavel": responsible_name(group),
                "telefones": "",
                "status_pendencia": "sem_canal_envio",
                "observacoes": "Fora do seed e do disparo ate haver telefone de envio.",
            })
            continue
        if len(responsible) == 0:
            errors.append(f"{group_id}: no family responsible")
        if len(responsible) > 1:
            errors.append(f"{group_id}: multiple family responsibles")

        for row in group:
            row_observations = "; ".join(
                item for item in [row.get("observacoes", ""), row.get("status_revisao", "")] if item
            )
            seed_rows.append({
                "id_familia": group_id,
                "nome_familia": family_name(group),
                "telefone_whatsapp": row.get("telefone_whatsapp", ""),
                "nome_convidado": row["nome_convidado"],
                "e_crianca": row["e_crianca"],
                "e_responsavel_familia": row["e_responsavel_familia"],
                "email": "",
                "observacoes": row_observations,
            })

        code = codes.get(group_id, "")
        link = f"{host}/rsvp/{code}" if code else ""
        status = "pendente_envio" if link else "pendente_seed_rsvp"
        for phone in phones:
            dispatch_rows.append({
                "id_familia": group_id,
                "nome_familia": family_name(group),
                "telefone_whatsapp": phone,
                "nomes_grupo": " | ".join(row["nome_convidado"] for row in group),
                "responsavel": responsible_name(group),
                "codigo_rsvp": code,
                "link_rsvp": link,
                "status_envio": status,
                "observacoes": "" if link else "Aguardando seed/commit para preencher link_rsvp real.",
            })

    for group_id, group in extra_groups.items():
        pending_rows.append({
            "id_familia": group_id,
            "nome_familia": family_name(group),
            "origem": "LISTA EXTRA",
            "nomes_grupo": " | ".join(row["nome_convidado"] for row in group),
            "responsavel": responsible_name(group),
            "telefones": " | ".join(group_phones(group)),
            "status_pendencia": "lote_pendente_contato;fora_primeiro_disparo",
            "observacoes": "Nao entra no primeiro lote.",
        })

    return seed_rows, dispatch_rows, pending_rows, errors


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=SOURCE_XLSX)
    parser.add_argument("--out-dir", type=Path, default=OUT_DIR)
    parser.add_argument("--host", default=HOST)
    parser.add_argument("--codes-file", type=Path, default=None)
    args = parser.parse_args()

    rows = read_xlsx(args.source)
    codes = load_codes(args.codes_file)
    seed_rows, dispatch_rows, pending_rows, errors = build_outputs(rows, codes, args.host.rstrip("/"))
    if errors:
        for error in errors:
            print(f"ERROR: {error}")
        raise SystemExit(1)

    seed_file = args.out_dir / "planilha-convidados-seed.csv"
    dispatch_file = args.out_dir / "disparo-primeiro-lote.csv"
    pending_file = args.out_dir / "grupos-pendentes-contato.csv"
    summary_file = args.out_dir / "rsvp-operacional-resumo.json"

    write_csv(seed_file, seed_rows, SEED_COLUMNS)
    write_csv(dispatch_file, dispatch_rows, DISPATCH_COLUMNS)
    write_csv(pending_file, pending_rows, PENDING_COLUMNS)

    source_counter = Counter(row.get("origem", "") for row in rows)
    summary = {
        "source_xlsx": str(args.source),
        "seed_csv": str(seed_file),
        "dispatch_csv": str(dispatch_file),
        "pending_csv": str(pending_file),
        "total_rows_source": len(rows),
        "source_rows_by_origem": dict(source_counter),
        "first_lot_guest_rows_seeded": len(seed_rows),
        "first_lot_families": len({row["id_familia"] for row in seed_rows}),
        "first_lot_dispatch_channels": len(dispatch_rows),
        "pending_groups": len(pending_rows),
        "links_from_codes_file": bool(codes),
    }
    summary_file.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
